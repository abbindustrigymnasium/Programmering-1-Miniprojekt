import openpyxl
wb= openpyxl.load_workbook('mello.xlsx')
sheet= wb.active


# print(sheet['A1'].value)
# print(sheet.cell(row=1, column=1).value)

# värden=[]
# for i in range(2,10):
#         värden.append(i)
# print(värden)

# värden=[ "Värde "+str(i) for i in range(2,10) if i%2 is 0]
# värden=[ "Värde "+str(i) for i in range(2,10) if i%2 is 0 if i is not 8]
# print(värden)

# artist=[]
# for i in range(2,sheet.max_row):
#     artist.append(sheet.cell(row=i, column=2).value)
# print(len(artist))

# artists=[ sheet.cell(row=i, column=2).value for i in range(2,sheet.max_row)]



# Nano = {'Namn': sheet.cell(row=2, column=2).value, 'Sång': sheet.cell(row=2, column=3).value, 'Poäng': sheet.cell(row=2, column=6).value, 'Röst': sheet.cell(row=2, column=5).value}

artists=[{"Namn":sheet.cell(row=i, column=2).value,
        "Sång":sheet.cell(row=i, column=3).value,
        "Poäng":sheet.cell(row=i, column=6).value,
        "Röst":sheet.cell(row=i, column=5).value
} for i in range(2,sheet.max_row)
    if not sheet.cell(row=i, column=8).value.startswith("Utsl") ]

print(len(artists))

Vidare=[{"Namn":sheet.cell(row=i, column=2).value,
        "Sång":sheet.cell(row=i, column=3).value,
        "Poäng":sheet.cell(row=i, column=6).value,
        "Röst":sheet.cell(row=i, column=5).value
} for i in range(2,sheet.max_row)
    if not sheet.cell(row=i, column=8).value.startswith("Utsl") ]

print(len(Vidare))



[print(artist["Namn"]," är med i år!") for artist in artists]
[print(artist["Namn"],"var en av de",len(Vidare)," som gick vidare med låten",artist["Sång"]) for artist in Vidare]
# print(Vidare)

Vidare = sorted(Vidare, key=lambda k: k['Poäng'], reverse=True) 
# OVidare= sorted(OVidare, key=lambda x: x['Poäng'], reverse=True)
# print("Dessa borde kommit i topp:")
Vidare=Vidare[0:10]
[print(artist["Namn"]," ",artist["Poäng"]," ",artist["Röst"]) for artist in Vidare[:3]]

# # [print(artist["Namn"]," ",type(artist["Poäng"])) for artist in Vidare]
# # print("\n Inte vidare:")
# # [print(artist["Namn"]," ",artist["Poäng"]) for artist in OVidare]