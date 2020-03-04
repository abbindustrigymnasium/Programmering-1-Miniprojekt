import openpyxl
wb= openpyxl.load_workbook('Farmen.xlsx')
# sheet= wb.active

# print(wb.get_sheet_names())
# Deltagar_sheet= wb.get_sheet_by_name('Deltagare')
# artists=[{"Namn":sheet.cell(row=i, column=2).value,
#         "Sång":sheet.cell(row=i, column=3).value,
#         "Poäng":sheet.cell(row=i, column=6).value,
#         "Röst":sheet.cell(row=i, column=5).value
# } for i in range(2,sheet.max_row) ]

# print(artists)